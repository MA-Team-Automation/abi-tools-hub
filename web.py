"""ABI Tools Hub Web 控制台。

纯标准库实现：http.server.ThreadingHTTPServer 提供 JSON API 与单页前端，
前端页面为 static/index.html（内嵌 CSS/JS，无框架无构建）。
所有耗时操作（更新 / 搭建环境 / 全量刷新）均在后台线程执行，不阻塞 HTTP 响应。
"""

import json
import os
import shutil
import subprocess
import sys
import threading
import time
import webbrowser
from collections import deque
from concurrent.futures import ThreadPoolExecutor
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import core
import jobs

ROOT = Path(__file__).resolve().parent
INDEX_HTML = ROOT / "static" / "index.html"

# 内存日志环形缓冲：保留最近 500 行，每行带单调递增序号供前端增量拉取
MAX_LOG_LINES = 500
_log_lock = threading.Lock()
_log_lines = deque(maxlen=MAX_LOG_LINES)  # 元素为 (seq, text)
_log_seq = 0

# 各工具正在执行的操作（"update" / "setup"），前端据此禁用对应按钮
_state_lock = threading.Lock()
_running = {}
_refresh_running = False

# 快速状态缓存：启动时后台线程并行 check_status（git fetch 较慢），
# 就绪前 /api/status 只返回本地瞬时信息（exists/env_ready，不走网络）
_status_lock = threading.Lock()
_status_cache = None  # dict: tool_id -> check_status 结果；None 表示尚未就绪
_status_checking = False

# 前期准备阶段检查结果缓存：启动时后台线程执行 core.check_prepare()。
# 就绪前 /api/prepare 返回 None（前端显示「准备检查中…」）
_prepare_lock = threading.Lock()
_prepare_cache = None  # dict: check_prepare() 结果；None 表示尚未就绪


def _check_prepare_background() -> None:
    """后台线程：执行一次前期准备阶段的环境检查并缓存结果。"""
    global _prepare_cache
    try:
        with _prepare_lock:
            _prepare_cache = core.check_prepare()
        log("前期准备检查完成。")
    except Exception as exc:
        log(f"错误：前期准备检查失败：{exc}")

# Hub 自更新状态：behind 为最近一次检查到的落后提交数；
# updated=True 表示本次运行期间拉到了新版本（需重启生效，重启前保持 True）
_hub_status = {"behind": None, "updated": False}

# 仓库访问权限/认证失败的典型日志特征（小写匹配）：
# GitHub 私有仓库无权限或未登录时，clone/fetch/pull 的常见报错
_ACCESS_ERROR_PATTERNS = (
    "repository not found",    # GitHub 私有仓库无权限（或未登录）时的 404
    "remote: not found",       # 同上，remote 前缀形式
    "authentication failed",   # 凭据无效/过期
    "could not read username", # 无交互终端下询问用户名失败（未配置凭据）
    "403",                     # HTTP 403 Forbidden
)

# 会话级权限错误标记：一旦置 True 本会话保持（语义同 hub.updated）
_access_error = False

# 全量刷新进度：total 固定为 6（1 个 Hub 自更新 + 5 个工具）；
# current 为正在进行的单元显示名列表（并行时可能有多个），供前端展示真实进度
_refresh_progress = {"active": False, "done": 0, "total": 6, "current": []}


def _is_access_error(text: str) -> bool:
    """判断日志文本是否包含仓库访问权限/认证失败特征（大小写不敏感）。"""
    lowered = text.lower()
    return any(p in lowered for p in _ACCESS_ERROR_PATTERNS)


def _mark_access_error(line: str) -> None:
    """检查一行日志，命中权限特征则置起会话级 access_error 标记。"""
    global _access_error
    if _access_error:
        return
    if _is_access_error(line):
        with _state_lock:
            _access_error = True
        log("警告：检测到仓库访问权限问题（可能需要申请组织权限或完成 GitHub 登录）")

# 进程启动时的原始 argv，/api/restart 原地重启时原样复用
_LAUNCH_ARGV = sys.argv[:]
# serve() 启动后保存服务器实例，重启前先释放端口
_server = None


def log(message: str) -> None:
    """追加一行带时间戳的日志到内存环形缓冲。"""
    global _log_seq
    line = f"[{time.strftime('%H:%M:%S')}] {message}"
    with _log_lock:
        _log_seq += 1
        _log_lines.append((_log_seq, line))


def _find_tool(tool_id: str):
    """按 id 在 tools.json 清单中查找工具，找不到返回 None。"""
    for tool in core.load_tools():
        if tool["id"] == tool_id:
            return tool
    return None


def _find_action(tool: dict, action_id: str):
    """在工具的 actions 中按 id 查找操作，找不到返回 None。"""
    for action in tool.get("actions", []):
        if action["id"] == action_id:
            return action
    return None


def _tools_payload() -> dict:
    """组装 /api/tools 响应：工具展示字段 + actions 定义，供前端渲染操作表单。"""
    tools = []
    for tool in core.load_tools():
        tools.append({
            "id": tool["id"],
            "name": tool["name"],
            "display": tool["display"],
            "description": tool["description"],
            "branch": tool["branch"],
            "entry": tool["entry"],
            "actions": tool.get("actions", []),
        })
    return {"tools": tools}


def _run_action(tool: dict, action: str, func) -> None:
    """后台线程执行耗时操作，写日志并维护 running 状态。"""
    action_name = {"update": "更新", "setup": "搭建环境"}[action]
    with _state_lock:
        _running[tool["id"]] = action
    log(f"[{tool['id']}] 开始{action_name}...")
    try:
        for line in str(func(tool)).splitlines():
            log(line)
        log(f"[{tool['id']}] {action_name}完成")
    except Exception as exc:
        log(f"[{tool['id']}] 错误：{action_name}失败：{exc}")
    finally:
        with _state_lock:
            _running.pop(tool["id"], None)


def _progress_begin(total: int) -> None:
    """开始一轮刷新：重置进度（各单元开跑后自行登记到 current）。"""
    with _state_lock:
        _refresh_progress.update({"active": True, "done": 0, "total": total,
                                  "current": []})


def _progress_start(name: str) -> None:
    """登记一个单元进入进行中（可在工作线程并发调用）。"""
    with _state_lock:
        _refresh_progress["current"].append(name)


def _progress_done(name: str) -> None:
    """登记一个单元完成：done+1 并移出进行中列表。"""
    with _state_lock:
        _refresh_progress["done"] += 1
        try:
            _refresh_progress["current"].remove(name)
        except ValueError:
            pass


def _progress_end() -> None:
    """整轮结束。"""
    with _state_lock:
        _refresh_progress.update({"active": False, "current": []})


def _refresh_all_background() -> None:
    """后台线程：单轮并行流程——Hub 自更新与 5 个工具共 6 个单元放进同一个

    ThreadPoolExecutor(max_workers=6) 同时开跑，墙钟时间 ≈ 最慢单个单元；
    结束后用本轮 fetch 后的本地 git 信息直接构建状态缓存（不再来一轮 fetch）。
    """
    global _refresh_running
    with _state_lock:
        if _refresh_running:
            return
        _refresh_running = True
    tools = core.load_tools()
    hub_unit = "ABI 工具箱"
    _progress_begin(total=1 + len(tools))
    log("开始自动检查更新并搭建环境（6 个单元并行，git fetch 可能需要几分钟）...")

    def _hub_unit() -> None:
        """并行单元：Hub 自身更新（含 main 分支/干净工作区/ff-only 保护）。"""
        _progress_start(hub_unit)
        try:
            result = core.self_update()
            for line in result["log"].splitlines():
                log(line)
                _mark_access_error(line)
            with _state_lock:
                _hub_status["behind"] = result["behind"]
                if result["updated"]:
                    _hub_status["updated"] = True
        except Exception as exc:
            log(f"错误：Hub 自更新检查失败：{exc}")
        finally:
            _progress_done(hub_unit)

    def _tool_unit(tool: dict) -> None:
        """并行单元：单个工具的 ensure_repo → update_tool → 按需 setup_env。"""
        _progress_start(tool["display"])
        try:
            for line in core.refresh_tool(tool):
                log(line)
                _mark_access_error(line)
        except Exception as exc:
            log(f"[{tool['id']}] 错误：更新失败：{exc}")
        finally:
            _progress_done(tool["display"])

    try:
        with ThreadPoolExecutor(max_workers=6) as pool:
            futures = [pool.submit(_hub_unit)]
            futures.extend(pool.submit(_tool_unit, tool) for tool in tools)
            for future in futures:
                future.result()  # 单元内部已捕获异常，此处仅等待全部完成
        log("自动更新与环境检查完成。")
    except Exception as exc:
        log(f"错误：自动更新失败：{exc}")
    finally:
        with _state_lock:
            _refresh_running = False
        _progress_end()
    # 用本轮 fetch/pull 后的本地 git 信息构建状态缓存（pull 后 behind 自然为 0，
    # 无需再来一轮 fetch；纯本地命令，毫秒级完成）
    _build_status_cache_local()


def _build_status_cache_local() -> None:
    """并行执行 local_status（不 fetch）填充状态缓存。"""
    global _status_cache
    try:
        results = {}
        with ThreadPoolExecutor(max_workers=5) as pool:
            for status in pool.map(core.local_status, core.load_tools()):
                results[status["id"]] = status
        with _status_lock:
            _status_cache = results
        log("状态缓存已就绪（基于本轮更新结果）。")
    except Exception as exc:
        log(f"错误：状态缓存构建失败：{exc}")


def _do_restart() -> None:
    """延迟执行的原地重启：先释放端口，再用原始 argv 拉起新进程。"""
    log("正在重启 Hub 服务...")
    try:
        if _server is not None:
            _server.server_close()
    except Exception:
        pass
    argv = [sys.executable, *_LAUNCH_ARGV]
    if sys.platform == "win32":
        # Windows 的 os.execv 不处理可执行路径中的空格
        # （C:\Program Files\... 会在空格处截断导致新进程启动失败），
        # 改用 Popen（正确拼接命令行并继承同一控制台）后退出当前进程，
        # 效果等同原地重启。
        subprocess.Popen(argv)
        os._exit(0)
    os.execv(sys.executable, argv)


def _quick_status(tool: dict) -> dict:
    """本地瞬时状态：只看文件系统（exists/env_ready），不执行 git 网络操作。"""
    path = Path(tool["path"])
    exists = (path / ".git").is_dir()
    return {
        "id": tool["id"],
        "exists": exists,
        "branch": None,
        "behind": None,
        "ahead": None,
        "last_commit": None,
        "env_ready": exists and (path / ".venv").is_dir(),
        "update_available": None,
    }


def _check_status_background() -> None:
    """后台线程：并行 check_status 全部工具并填充缓存（总耗时接近最慢单个 fetch）。"""
    global _status_cache, _status_checking
    with _status_lock:
        if _status_checking:
            return
        _status_checking = True
    try:
        results = {}
        with ThreadPoolExecutor(max_workers=5) as pool:
            for status in pool.map(core.check_status, core.load_tools()):
                results[status["id"]] = status
        with _status_lock:
            _status_cache = results
        log("后台状态检查完成。")
    except Exception as exc:
        log(f"错误：后台状态检查失败：{exc}")
    finally:
        with _status_lock:
            _status_checking = False


def _status_payload() -> dict:
    """组装 /api/status 响应。

    缓存就绪后返回完整 check_status 结果；未就绪时返回本地瞬时信息
    （exists/env_ready，git 字段为 None）。refreshing 表示后台状态检查
    或全量更新是否仍在进行。
    """
    with _state_lock:
        running = dict(_running)
        refresh_running = _refresh_running
        hub = dict(_hub_status)
        access_error = _access_error
        refresh = dict(_refresh_progress)
        refresh["current"] = list(_refresh_progress["current"])
    with _status_lock:
        cache = dict(_status_cache) if _status_cache is not None else None
        checking = _status_checking
    tools = []
    for tool in core.load_tools():
        if cache is not None and tool["id"] in cache:
            status = dict(cache[tool["id"]])
        else:
            status = _quick_status(tool)
        status["display"] = tool["display"]
        status["description"] = tool["description"]
        status["running"] = running.get(tool["id"])
        tools.append(status)
    return {
        "tools": tools,
        "refresh_running": refresh_running,
        "refreshing": refresh_running or checking,
        "refresh": refresh,
        "hub": hub,
        "access_error": access_error,
    }


def _prepare_payload() -> dict:
    """组装 /api/prepare 响应：前期准备检查结果（未就绪时 ready=False）。"""
    with _prepare_lock:
        cache = dict(_prepare_cache) if _prepare_cache is not None else None
    if cache is None:
        return {"ready": False, "checks": None}
    failed = [c for c in cache.values() if not c["ok"]]
    return {"ready": True, "checks": cache, "passed": len(failed) == 0,
            "failed_count": len(failed)}


def _logs_payload(since: int) -> dict:
    """组装 /api/logs 响应：序号大于 since 的日志行 + 当前最大序号。"""
    with _log_lock:
        lines = [text for seq, text in _log_lines if seq > since]
        last = _log_seq
    return {"lines": lines, "last": last}


# /api/inspect-file 在该工具 .venv 中执行的驱动代码：
# 用 openpyxl 只读模式列出 Sheet 名或指定 Sheet 的首行表头，结果以标记行输出 JSON
_INSPECT_MARKER = "__INSPECT_RESULT__"
_INSPECT_CODE = r"""
import json
import sys

from openpyxl import load_workbook

path, what = sys.argv[1], sys.argv[2]
sheet = sys.argv[3] if len(sys.argv) > 3 else ""

wb = load_workbook(path, read_only=True)
try:
    if what == "sheets":
        result = list(wb.sheetnames)
    else:
        name = sheet or wb.sheetnames[0]
        if name not in wb.sheetnames:
            raise ValueError(f"Sheet 不存在：{name}")
        row = next(wb[name].iter_rows(min_row=1, max_row=1, values_only=True), ())
        result = [str(c).strip() for c in row if c is not None and str(c).strip()]
    print("__INSPECT_RESULT__" + json.dumps(result, ensure_ascii=False))
finally:
    wb.close()
"""


def _inspect_file(body: dict) -> tuple:
    """读取 Excel 文件的 Sheet 列表或首行表头（供前端动态下拉选项）。

    在对应工具的 .venv 里同步执行 openpyxl（read_only 模式），超时 120 秒。
    返回 (payload, http_status)；所有失败路径均为中文错误信息。
    """
    tool_id = str(body.get("tool_id") or "")
    what = str(body.get("what") or "")
    path_text = str(body.get("path") or "").strip().strip('"').strip("'")
    sheet = str(body.get("sheet") or "").strip()

    tool = _find_tool(tool_id)
    if tool is None:
        return {"ok": False, "error": f"未知工具：{tool_id}"}, 404
    if what not in ("sheets", "columns"):
        return {"ok": False, "error": "what 必须是 sheets 或 columns"}, 400
    if not path_text:
        return {"ok": False, "error": "请先填写文件路径"}, 400
    file_path = Path(path_text)
    if not file_path.is_file():
        return {"ok": False, "error": f"文件不存在：{path_text}"}, 400
    if file_path.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"ok": False, "error": "只能读取 .xlsx 格式的 Excel 文件"}, 400
    if shutil.which("uv") is None:
        return {"ok": False, "error": "未找到 uv，无法读取文件"}, 500

    try:
        r = subprocess.run(
            ["uv", "run", "python", "-X", "utf8", "-c", _INSPECT_CODE,
             str(file_path), what, sheet],
            cwd=tool["path"],
            capture_output=True, text=True,
            encoding="utf-8", errors="replace",
            timeout=120,
        )
    except subprocess.TimeoutExpired:
        return {"ok": False, "error": "读取超时（120 秒），文件可能过大或被占用"}, 500

    if r.returncode != 0:
        detail = ""
        for line in (r.stderr or "").strip().splitlines():
            line = line.strip()
            if line and "Warning" not in line:
                detail = line  # 取最后一行有效报错
        return {"ok": False, "error": f"读取 Excel 失败：{detail or '未知错误'}"}, 400

    for line in (r.stdout or "").splitlines():
        if line.startswith(_INSPECT_MARKER):
            try:
                options = json.loads(line[len(_INSPECT_MARKER):])
            except json.JSONDecodeError:
                break
            return {"ok": True, "options": options}, 200
    return {"ok": False, "error": "读取 Excel 失败：未获得有效结果"}, 500


class HubRequestHandler(BaseHTTPRequestHandler):
    """路由处理：静态页面 + JSON API。"""

    server_version = "AbiToolsHub/1.0"

    # 默认日志打到 stderr 太吵，静默处理
    def log_message(self, format, *args):
        pass

    def _send_json(self, payload: dict, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_index(self) -> None:
        body = INDEX_HTML.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        parts = [p for p in parsed.path.split("/") if p]

        if parsed.path in ("/", "/index.html"):
            self._send_index()
        elif parsed.path == "/api/status":
            self._send_json(_status_payload())
        elif parsed.path == "/api/prepare":
            self._send_json(_prepare_payload())
        elif parsed.path == "/api/tools":
            self._send_json(_tools_payload())
        elif parsed.path == "/api/jobs":
            self._send_json({"jobs": jobs.get_jobs()})
        elif len(parts) == 3 and parts[0] == "api" and parts[1] == "job":
            # /api/job/<job_id>?since=N —— 任务增量日志轮询
            qs = parse_qs(parsed.query)
            try:
                since = int(qs.get("since", ["0"])[0])
            except ValueError:
                since = 0
            result = jobs.get_job_log(parts[2], since=since)
            if result is None:
                self._send_json({"error": f"任务不存在：{parts[2]}"}, status=404)
            else:
                self._send_json(result)
        elif parsed.path == "/api/logs":
            qs = parse_qs(parsed.query)
            try:
                since = int(qs.get("since", ["0"])[0])
            except ValueError:
                since = 0
            self._send_json(_logs_payload(since))
        else:
            self._send_json({"error": "not found"}, status=404)

    def _read_json_body(self):
        """读取并解析请求体 JSON，失败返回 None（调用方负责回 400）。"""
        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError:
            return None
        try:
            return json.loads(self.rfile.read(length).decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return None

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        parts = [p for p in parsed.path.split("/") if p]

        if parsed.path == "/api/refresh-all":
            threading.Thread(target=_refresh_all_background, daemon=True).start()
            self._send_json({"ok": True, "message": "已开始全量更新与环境搭建（后台执行）"})
            return

        # /api/inspect-file —— 读取 Excel 的 Sheet 列表 / 首行表头，供动态下拉选项
        if parsed.path == "/api/inspect-file":
            body = self._read_json_body()
            if body is None or not isinstance(body, dict):
                self._send_json({"ok": False, "error": "请求体必须是 JSON 对象"}, status=400)
                return
            payload, status = _inspect_file(body)
            self._send_json(payload, status=status)
            return

        # /api/restart —— 原地重启 Hub 进程（有任务运行时拒绝）
        if parsed.path == "/api/restart":
            with _state_lock:
                busy = _refresh_running or bool(_running)
            running_jobs = [j for j in jobs.get_jobs() if j["status"] == "running"]
            if busy or running_jobs:
                self._send_json(
                    {"ok": False, "error": "有任务正在运行，请稍后重启"}, status=409)
                return
            self._send_json({"ok": True, "message": "正在重启…"})
            timer = threading.Timer(0.5, _do_restart)
            timer.daemon = False  # 保证 0.5 秒后能执行到 execv
            timer.start()
            return

        # /api/run/<tool_id>/<action_id> —— 启动工具操作任务，返回 {"job_id": "..."}
        if len(parts) == 4 and parts[0] == "api" and parts[1] == "run":
            tool_id, action_id = parts[2], parts[3]
            tool = _find_tool(tool_id)
            if tool is None:
                self._send_json({"ok": False, "error": f"未知工具：{tool_id}"}, status=404)
                return
            action = _find_action(tool, action_id)
            if action is None:
                self._send_json({"ok": False, "error": f"未知操作：{action_id}"}, status=404)
                return
            body = self._read_json_body()
            if body is None or not isinstance(body, dict):
                self._send_json({"ok": False, "error": "请求体必须是 JSON 对象"}, status=400)
                return
            try:
                job_id = jobs.start_job(tool, action, body.get("params") or {})
            except ValueError as exc:
                self._send_json({"ok": False, "error": str(exc)}, status=400)
                return
            log(f"[{tool_id}] 已开始「{action['label']}」（{job_id}）")
            self._send_json({"job_id": job_id})
            return

        # /api/<update|setup|launch>/<tool_id>
        if len(parts) == 3 and parts[0] == "api" and parts[1] in ("update", "setup", "launch"):
            action, tool_id = parts[1], parts[2]
            tool = _find_tool(tool_id)
            if tool is None:
                self._send_json({"ok": False, "error": f"未知工具：{tool_id}"}, status=404)
                return

            if action == "launch":
                # launch 本身很快（os.startfile / Popen 立即返回），同步执行
                try:
                    core.launch_tool(tool)
                    log(f"[{tool_id}] 已启动「{tool['display']}」")
                    self._send_json({"ok": True, "message": f"已启动「{tool['display']}」"})
                except Exception as exc:
                    log(f"[{tool_id}] 错误：启动失败：{exc}")
                    self._send_json({"ok": False, "error": str(exc)}, status=500)
                return

            # update / setup 耗时较长，放后台线程执行
            func = core.update_tool if action == "update" else core.setup_env
            with _state_lock:
                if tool_id in _running:
                    self._send_json({"ok": False, "error": "该工具已有操作在进行中"}, status=409)
                    return
            threading.Thread(target=_run_action, args=(tool, action, func), daemon=True).start()
            action_name = "更新" if action == "update" else "搭建环境"
            self._send_json({"ok": True, "message": f"已开始{action_name}「{tool['display']}」（后台执行）"})
            return

        self._send_json({"error": "not found"}, status=404)


def serve(port: int, open_browser: bool = True, auto_refresh: bool = True) -> None:
    """启动 Web 控制台。

    auto_refresh 为 True 时开 daemon 线程执行单轮刷新（自更新 + 并行更新工具 +
    构建状态缓存）；为 False（--no-update）时只开后台线程并行 check_status
    填充缓存（含 fetch）。两条路径互斥，避免对同一批仓库重复 fetch。
    open_browser 为 True 时延迟 1 秒用默认浏览器打开控制台页面。
    """
    if auto_refresh:
        threading.Thread(target=_refresh_all_background, daemon=True).start()
    else:
        threading.Thread(target=_check_status_background, daemon=True).start()
    # 前期准备阶段环境检查（Python/Git/uv/GitHub/组织），后台并行执行
    threading.Thread(target=_check_prepare_background, daemon=True).start()

    global _server
    server = ThreadingHTTPServer(("127.0.0.1", port), HubRequestHandler)
    _server = server
    url = f"http://127.0.0.1:{port}"
    if open_browser:
        threading.Timer(1.0, webbrowser.open, args=(url,)).start()
    log(f"Web 控制台已启动：{url}")
    print(f"Web 控制台已启动：{url}")
    print("按 Ctrl+C 停止。")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n正在停止 Web 控制台...")
    finally:
        server.server_close()
